"""
    Name: easyHtmlTable
    Author: Ming Yin Kenneth Chau
    Description: Create easy html table from excel file
"""

import openpyxl

# Define the easyHtmlTable function
def easyHtmlTable(filename: str, sheetname: str, containHyper:int = None, firstColumn = 1, firstRow = 1, headerRow = True, htmlTableClass:str = None, htmlTableId:str = None):
    workBook = openpyxl.load_workbook(filename)
    workSheet = workBook[sheetname]

    # Now we create a list for headers and a list of list for the body
    header = []
    body = []
    htmlBody = ""

    # Get the size of the table
    maxColumn = workSheet.max_column
    maxRow = workSheet.max_row

    # Grab the header and store it in the list
    if headerRow == True: 
        for item in range(firstColumn, maxColumn):
            header.append(workSheet.cell(firstRow, column = item).value)
    
    # get the body and store them in the list of list
    for rowNo in range(firstRow + 1, maxRow +1):
        rowElement = []
        for columnNo in range(1, maxColumn):
            if containHyper == None:
                rowElement.append(workSheet.cell(row = rowNo, column = columnNo).value)
            elif (columnNo == containHyper):
                HyperList = []
                target = []
                try:
                    target.append(workSheet.cell(row = rowNo, column = columnNo).hyperlink.target)
                    target.append(workSheet.cell(row = rowNo, column = columnNo).value)
                # Due to openpyxl bug in reading hyperlink, use string formatting
                except:
                    target = workSheet.cell(row = rowNo, column = columnNo).value
                    target = target[target.find("("):][1:-1].replace('"','').split(',')
                HyperList.append(target[1].strip())
                HyperList.append(target[0].strip())
                rowElement.append(HyperList)
            else:
                rowElement.append(workSheet.cell(row = rowNo, column = columnNo).value)
        body.append(rowElement)
    

    # Assemble the html header
    HeaderElements = ''
    for item in header:
        HeaderElements = HeaderElements + '\n' + "<th>{}</th>".format(item) 
    header = "<thead>\n<tr>{}</tr>\n</thead>".format(HeaderElements)
    
    # Assemble the body
    BodyElements = ''
    for row in body:
        rowElement = ''
        Entry = ''
        for element in row:
            if type(element) is list:
                rowElement = '<td><a href="{}">{}</a></td>'.format(element[1], element[0])
            else:
                rowElement = "<td>{}</td>".format(element)
            Entry =  Entry + rowElement + '\n'
        BodyElements = BodyElements + '\n<tr>\n' + Entry + '\n</tr>'
    body = '<tbody>\n{}\n</tbody>'.format(BodyElements)

    # Assemble the table tag
    tableTag = 'table'
    if htmlTableClass != None:
        tableTag = tableTag + ' class="{}"'.format(htmlTableClass)
    if htmlTableId != None:
        tableTag = tableTag + ' id="{}"'.format(htmlTableId)
    tableTag = "<{}>\n{{}}\n</table>".format(tableTag)

    # Assemble the html
    htmlBody = tableTag.format(header + '\n' + body)

    # Write the output html file
    with open('output.html','w', encoding = 'utf-8-sig') as output:
        output.write(htmlBody)




def main():
    easyHtmlTable('SAP Ariba_resource guides (3).xlsx', 'Sheet1', containHyper = 1, htmlTableId= 'SAPAribaResourceGuide', htmlTableClass='wb-tables table')

if __name__ == "__main__":
    main()